# coding=utf-8
#
# Author：ECNU 虾饺同学
#

import os
import io
import sys
import openpyxl

#白名单和黑名单 格式：'关键词1','关键词2'
allowed_list = []

block_list = [
            '收回',
            '收到',
            '归还',
            '下属公司',
            '参股公司'
            ]

def xls_bak():
    os.system('cp result.xlsx result_old.xlsx')


def isExist(file):
    filepath = './txt/' + file[:-4] + '.txt'
    if os.path.exists(filepath):
        return True
    else:
        return False


def main(argv):
    pdf_dir = './pdf/'
    txt_dir = './txt/'
    
    wb = openpyxl.load_workbook('result.xlsx')
    ws = wb.active

    for file in os.listdir(pdf_dir):
        
        file_prefix = file[:-4]
        pdf_path = pdf_dir + file_prefix + '.pdf'
        txt_path = txt_dir + file_prefix + '.txt'
        
        if file.lower().endswith('.pdf'):

            allowed = True
            if len(allowed_list) != 0:
               allowed = file in allowed_list

            if len(block_list) != 0:
                for item in block_list:
                    if item in file:
                        allowed = False
                        break

            if allowed == False:
                continue

            if isExist(file):
                continue
            
            cmd = 'node extract_text.js ' + pdf_path + " " + txt_path
            os.system(cmd)

        elif file.lower().endswith('.txt'):
            if isExist(file):
                continue
            
            cmd = 'cp ' + pdf_dir + file + " " + txt_dir
            os.system(cmd)

        xls_bak()
            
        try:
            with io.open(txt_path,encoding='UTF-8') as txt_file:
                lines = txt_file.readlines()
                content = ""
                for line in lines:
                    line = line.strip()
                    if len(line) != 0:
                        content += line[:len(line)-1]

            code = file_prefix[0:6]
            date = file_prefix[7:15]
            title = file_prefix[16:len(file_prefix)]

            insert_row = ws.max_row + 1
            ws.cell(row = insert_row, column = 1, value = code)
            ws.cell(row = insert_row, column = 2, value = date)
            ws.cell(row = insert_row, column = 3, value = title)
            ws.cell(row = insert_row, column = 4, value = content)

            wb.save('result.xlsx')
        except:
            os.system('mv result_old.xlsx result.xlsx')
            insert_row = ws.max_row + 1
            ws.cell(row = insert_row, column = 1, value = code)
            ws.cell(row = insert_row, column = 2, value = date)
            ws.cell(row = insert_row, column = 3, value = "写入错误")
            wb.save('result.xlsx')


if __name__ == '__main__':
    main(sys.argv)
